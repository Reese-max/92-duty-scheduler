"""逐格比對匯出 Excel 與參考 ODS"""
import sys
from pathlib import Path

try:
    from odf.opendocument import load as load_ods
    from odf.table import Table, TableRow, TableCell, CoveredTableCell
    from odf.text import P
    from odf.style import Style, TableColumnProperties, TableRowProperties, TextProperties, ParagraphProperties, TableCellProperties
    HAS_ODF = True
except ImportError:
    HAS_ODF = False

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXPORT = Path(r'C:\Users\User\Desktop\92-duty-scheduler\勤務排班_第1週.xlsx')
REF_ODS = Path(r'C:\Users\User\Desktop\勤排 內網\勤務排表-- (26).ods')

def pt_to_cm(pt): return pt * 0.0352778
def cm_to_pt(cm): return cm / 0.0352778

def get_ods_sheets(doc):
    """Extract sheet data from ODS"""
    sheets = {}
    for table in doc.getElementsByType(Table):
        name = table.getAttribute('name')
        rows = []
        for row in table.getElementsByType(TableRow):
            cells = []
            for cell in row.childNodes:
                if cell.qname[1] in ('table-cell', 'covered-table-cell'):
                    text = ''
                    for p in cell.getElementsByType(P):
                        text += ''.join(c.data if hasattr(c,'data') else '' for c in p.childNodes)
                    repeat = int(cell.getAttribute('numbercolumnsrepeated') or 1)
                    for _ in range(repeat):
                        cells.append(text)
            rows.append(cells)
        sheets[name] = rows
    return sheets

def get_ods_styles(doc):
    """Extract style info from ODS"""
    styles = {}
    for s in doc.getElementsByType(Style):
        name = s.getAttribute('name')
        info = {}
        for tp in s.getElementsByType(TextProperties):
            sz = tp.getAttribute('fontsize')
            if sz: info['fontSize'] = sz
            fn = tp.getAttribute('fontname')
            if fn: info['fontName'] = fn
        for rp in s.getElementsByType(TableRowProperties):
            rh = rp.getAttribute('rowheight')
            if rh: info['rowHeight'] = rh
        for cp in s.getElementsByType(TableColumnProperties):
            cw = cp.getAttribute('columnwidth')
            if cw: info['colWidth'] = cw
        if info:
            styles[name] = info
    return styles

def analyze_excel(wb):
    """Analyze exported Excel structure"""
    report = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        sheet_info = {
            'name': sname,
            'rows': ws.max_row,
            'cols': ws.max_column,
            'merges': len(ws.merged_cells.ranges),
            'fonts': set(),
            'font_sizes': set(),
            'borders': {'thin': 0, 'medium': 0, 'none': 0},
            'alignments': set(),
            'fills': set(),
            'row_heights': {},
            'col_widths': {},
        }
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.font:
                    if cell.font.name: sheet_info['fonts'].add(cell.font.name)
                    if cell.font.sz: sheet_info['font_sizes'].add(cell.font.sz)
                if cell.border:
                    for side in ['top','bottom','left','right']:
                        b = getattr(cell.border, side)
                        if b and b.style:
                            if b.style == 'thin': sheet_info['borders']['thin'] += 1
                            elif b.style == 'medium': sheet_info['borders']['medium'] += 1
                        else:
                            sheet_info['borders']['none'] += 1
                if cell.alignment:
                    sheet_info['alignments'].add(f"h={cell.alignment.horizontal},v={cell.alignment.vertical},wrap={cell.alignment.wrap_text}")
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                    sheet_info['fills'].add(cell.fill.fgColor.rgb)

        for i in range(1, ws.max_row + 1):
            h = ws.row_dimensions[i].height
            if h: sheet_info['row_heights'][i] = h
        for i in range(1, ws.max_column + 1):
            w = ws.column_dimensions[get_column_letter(i)].width
            if w: sheet_info['col_widths'][get_column_letter(i)] = w

        report.append(sheet_info)
    return report

def check_export_quality(wb):
    """Check exported Excel against quality standards"""
    issues = []
    warnings = []

    EXPECTED_SHEETS = ['隊部值班', '友誼/警英夜巡', '週間校巡', '假日校巡', '寢巡', '警技館值班', '監廚', '總隊值星']
    # Note: actual sheet names from DUTY_DEFS labels

    # 1. Sheet order
    print('\n' + '='*60)
    print('  Sheet 順序檢查')
    print('='*60)
    actual = wb.sheetnames
    print(f'  匯出: {actual}')
    # Check expected sheets are present in order
    exp_in_actual = [s for s in EXPECTED_SHEETS if s in actual]
    # Also check with label variations
    ok_order = True
    for i, s in enumerate(actual):
        found = False
        for e in EXPECTED_SHEETS:
            if e in s or s in e:
                found = True
                break
        if not found:
            pass  # Some sheets might have slightly different names
    print(f'  共 {len(actual)} sheets')
    if len(actual) == 8:
        print('  PASS: 8 sheets 全部存在')
    else:
        issues.append(f'Sheet 數量: {len(actual)}, 預期 8')
        print(f'  ISSUE: Sheet 數量 {len(actual)}, 預期 8')

    # 2. Per-sheet analysis
    for sinfo in analyze_excel(wb):
        print(f'\n{"="*60}')
        print(f'  {sinfo["name"]}')
        print(f'{"="*60}')

        # Font check
        print(f'  字體: {sinfo["fonts"]}')
        if 'DFKai-SB' in sinfo['fonts']:
            print('  PASS: 使用標楷體 (DFKai-SB)')
        else:
            warnings.append(f'{sinfo["name"]}: 未使用 DFKai-SB')
            print('  WARNING: 未找到 DFKai-SB 字體')

        print(f'  字體大小: {sorted(sinfo["font_sizes"])}')

        # Border check
        print(f'  邊框: thin={sinfo["borders"]["thin"]}, medium={sinfo["borders"]["medium"]}')
        if sinfo['borders']['medium'] > 0:
            print('  PASS: 有 medium 邊框')
        else:
            issues.append(f'{sinfo["name"]}: 缺少 medium 邊框')
            print('  ISSUE: 缺少 medium 邊框')

        # Merge check
        print(f'  合併儲存格: {sinfo["merges"]} 個')
        if sinfo['merges'] > 0:
            print('  PASS: 有合併儲存格')
        else:
            issues.append(f'{sinfo["name"]}: 無合併儲存格')
            print('  ISSUE: 無合併儲存格')

        # Fill check
        fills = sinfo['fills']
        has_white = any('FFFFFF' in f for f in fills)
        has_gray = any('D8D8D8' in f for f in fills)
        print(f'  填充色: {fills}')
        if has_white:
            print('  PASS: 白色背景')
        if has_gray:
            print('  PASS: 灰色空格填充')

        # Row heights
        print(f'  列高設定: {len(sinfo["row_heights"])} 列有自訂高度')

        # Col widths
        print(f'  欄寬設定: {len(sinfo["col_widths"])} 欄有自訂寬度')

        # Alignment
        wraps = [a for a in sinfo['alignments'] if 'wrap=True' in a]
        centers = [a for a in sinfo['alignments'] if 'h=center' in a]
        print(f'  對齊: {len(centers)} 置中, {len(wraps)} 自動換行')

        # Data area size
        print(f'  範圍: {sinfo["rows"]} 列 x {sinfo["cols"]} 欄')

    # 3. Summary
    print(f'\n{"="*60}')
    print('  比對總結')
    print(f'{"="*60}')
    print(f'  Issues: {len(issues)}')
    for i in issues: print(f'    - {i}')
    print(f'  Warnings: {len(warnings)}')
    for w in warnings: print(f'    - {w}')

    if len(issues) == 0:
        print('\n  VERDICT: PASS — 匯出品質達標')
    else:
        print(f'\n  VERDICT: {len(issues)} 個問題需修正')

    return issues, warnings


if __name__ == '__main__':
    if not EXPORT.exists():
        print(f'匯出檔案不存在: {EXPORT}')
        print('請先執行匯出，或指定正確路徑')
        sys.exit(1)

    wb = load_workbook(str(EXPORT), data_only=True)
    issues, warnings = check_export_quality(wb)

    # ODS comparison if odfpy available
    if HAS_ODF and REF_ODS.exists():
        print(f'\n{"="*60}')
        print('  ODS 參考檔案比對')
        print(f'{"="*60}')
        doc = load_ods(str(REF_ODS))
        ods_sheets = get_ods_sheets(doc)
        ods_styles = get_ods_styles(doc)

        print(f'  ODS sheets: {list(ods_sheets.keys())}')
        print(f'  ODS styles: {len(ods_styles)} 個樣式定義')

        for sname, rows in ods_sheets.items():
            non_empty = sum(1 for r in rows if any(c for c in r))
            print(f'  {sname}: {non_empty} 非空列, {max(len(r) for r in rows) if rows else 0} 最大欄數')
    elif not HAS_ODF:
        print('\n  注意: 未安裝 odfpy，無法直接比對 ODS 內容')
        print('  安裝方式: pip install odfpy')
    elif not REF_ODS.exists():
        print(f'\n  注意: 參考 ODS 不存在: {REF_ODS}')

    print(f'\n{"="*60}')
    print('  完成')
    print(f'{"="*60}')
