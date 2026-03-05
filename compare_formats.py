#!/usr/bin/env python3
"""
Exhaustive comparison between exported Excel and reference ODS.
Compares all 8 duty types across 15 formatting categories.
Only prints DIFFERENCES.
"""

import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict

# =============================================================================
# Configuration
# =============================================================================

EXCEL_PATH = r"C:\Users\User\Desktop\92-duty-scheduler\勤務排班_第1週.xlsx"
ODS_PATH = r"C:\Users\User\Desktop\勤排\勤務排表-- (26).ods"

# Keyword map: Excel sheet name -> ODS keyword for matching
KEYWORD_MAP = {
    "隊部值班": "值班",
    "友誼警英夜巡": "夜巡",
    "週間校巡": "週間",
    "假日校巡": "校巡",
    "寢巡": "寢巡",
    "警技館值班": "警技館",
    "監廚": "監廚",
    "總隊值星": "總隊值星",
}

# ODS Namespaces
NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    "style": "urn:oasis:names:tc:opendocument:xmlns:style:1.0",
    "fo": "urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0",
    "number": "urn:oasis:names:tc:opendocument:xmlns:datastyle:1.0",
    "svg": "urn:oasis:names:tc:opendocument:xmlns:svg-compatible:1.0",
    "draw": "urn:oasis:names:tc:opendocument:xmlns:drawing:1.0",
}

# =============================================================================
# ODS Parsing Functions
# =============================================================================

def parse_ods(ods_path):
    """Parse ODS file and return structured data for all sheets."""
    with zipfile.ZipFile(ods_path, 'r') as z:
        content_xml = z.read('content.xml')

    root = ET.fromstring(content_xml)

    # Parse automatic styles
    styles = {}
    auto_styles = root.find('.//office:automatic-styles', NS)
    if auto_styles is not None:
        for style_elem in auto_styles:
            style_name = style_elem.get(f'{{{NS["style"]}}}name')
            if style_name:
                styles[style_name] = parse_style(style_elem)

    # Parse sheets
    sheets = {}
    body = root.find('.//office:body/office:spreadsheet', NS)
    if body is None:
        return sheets, styles

    for table in body.findall('table:table', NS):
        sheet_name = table.get(f'{{{NS["table"]}}}name')
        sheets[sheet_name] = parse_ods_sheet(table, styles)

    return sheets, styles


def parse_style(style_elem):
    """Parse a single ODS style element into a dict."""
    info = {
        'family': style_elem.get(f'{{{NS["style"]}}}family', ''),
        'parent': style_elem.get(f'{{{NS["style"]}}}parent-style-name', ''),
    }

    # Text properties
    tp = style_elem.find('style:text-properties', NS)
    if tp is not None:
        info['font-size'] = tp.get(f'{{{NS["fo"]}}}font-size', '')
        info['font-name'] = tp.get(f'{{{NS["style"]}}}font-name', '')
        info['font-weight'] = tp.get(f'{{{NS["fo"]}}}font-weight', '')
        info['font-style'] = tp.get(f'{{{NS["fo"]}}}font-style', '')
        info['font-size-asian'] = tp.get(f'{{{NS["style"]}}}font-size-asian', '')
        info['font-name-asian'] = tp.get(f'{{{NS["style"]}}}font-name-asian', '')
        info['color'] = tp.get(f'{{{NS["fo"]}}}color', '')

    # Table cell properties
    tcp = style_elem.find('style:table-cell-properties', NS)
    if tcp is not None:
        info['background-color'] = tcp.get(f'{{{NS["fo"]}}}background-color', '')
        info['border'] = tcp.get(f'{{{NS["fo"]}}}border', '')
        info['border-top'] = tcp.get(f'{{{NS["fo"]}}}border-top', '')
        info['border-bottom'] = tcp.get(f'{{{NS["fo"]}}}border-bottom', '')
        info['border-left'] = tcp.get(f'{{{NS["fo"]}}}border-left', '')
        info['border-right'] = tcp.get(f'{{{NS["fo"]}}}border-right', '')
        info['vertical-align'] = tcp.get(f'{{{NS["style"]}}}vertical-align', '')
        info['wrap-option'] = tcp.get(f'{{{NS["fo"]}}}wrap-option', '')
        info['shrink-to-fit'] = tcp.get(f'{{{NS["style"]}}}shrink-to-fit-properties', '')

    # Paragraph properties
    pp = style_elem.find('style:paragraph-properties', NS)
    if pp is not None:
        info['text-align'] = pp.get(f'{{{NS["fo"]}}}text-align', '')
        info['margin-left'] = pp.get(f'{{{NS["fo"]}}}margin-left', '')

    # Table row properties
    trp = style_elem.find('style:table-row-properties', NS)
    if trp is not None:
        info['row-height'] = trp.get(f'{{{NS["style"]}}}row-height', '')
        info['use-optimal-row-height'] = trp.get(f'{{{NS["style"]}}}use-optimal-row-height', '')

    # Table column properties
    tcp2 = style_elem.find('style:table-column-properties', NS)
    if tcp2 is not None:
        info['column-width'] = tcp2.get(f'{{{NS["style"]}}}column-width', '')

    return info


def get_cell_text(cell_elem):
    """Extract all text from a cell, joining text:p elements with newlines."""
    texts = []
    for p in cell_elem.findall('.//text:p', NS):
        parts = []
        if p.text:
            parts.append(p.text)
        for child in p:
            if child.text:
                parts.append(child.text)
            if child.tail:
                parts.append(child.tail)
        texts.append(''.join(parts))
    return '\n'.join(texts)


def parse_ods_sheet(table, styles):
    """Parse an ODS sheet into structured data."""
    sheet_data = {
        'columns': [],
        'rows': [],
        'merges': [],
    }

    # Parse columns
    for col in table.findall('table:table-column', NS):
        col_style = col.get(f'{{{NS["table"]}}}style-name', '')
        repeat = int(col.get(f'{{{NS["table"]}}}number-columns-repeated', '1'))
        col_info = {'style': col_style, 'repeat': repeat}
        if col_style in styles:
            col_info['width'] = styles[col_style].get('column-width', '')
        for _ in range(repeat):
            sheet_data['columns'].append(col_info)

    # Parse rows
    row_idx = 0
    for row_elem in table.findall('table:table-row', NS):
        row_style = row_elem.get(f'{{{NS["table"]}}}style-name', '')
        row_repeat = int(row_elem.get(f'{{{NS["table"]}}}number-rows-repeated', '1'))

        row_info = {
            'style': row_style,
            'height': styles.get(row_style, {}).get('row-height', ''),
            'cells': [],
        }

        col_idx = 0
        for cell in row_elem.findall('table:table-cell', NS):
            cell_style = cell.get(f'{{{NS["table"]}}}style-name', '')
            cell_repeat = int(cell.get(f'{{{NS["table"]}}}number-columns-repeated', '1'))
            col_span = int(cell.get(f'{{{NS["table"]}}}number-columns-spanned', '1'))
            row_span = int(cell.get(f'{{{NS["table"]}}}number-rows-spanned', '1'))
            text = get_cell_text(cell)
            value_type = cell.get(f'{{{NS["office"]}}}value-type', '')

            cell_info = {
                'style': cell_style,
                'text': text,
                'value_type': value_type,
                'col_span': col_span,
                'row_span': row_span,
                'col_idx': col_idx,
                'row_idx': row_idx,
            }
            if cell_style in styles:
                cell_info['style_data'] = styles[cell_style]

            if col_span > 1 or row_span > 1:
                sheet_data['merges'].append({
                    'row': row_idx, 'col': col_idx,
                    'row_span': row_span, 'col_span': col_span,
                })

            for r in range(cell_repeat):
                row_info['cells'].append(cell_info.copy())
                col_idx += 1

        # Only add the first instance for non-empty rows; skip large repeated empty rows
        if row_repeat > 50 and all(c['text'] == '' for c in row_info['cells']):
            row_idx += row_repeat
            continue

        for _ in range(min(row_repeat, 100)):  # cap at 100 repeated rows
            sheet_data['rows'].append(row_info)
            row_idx += 1

    return sheet_data


def find_best_ods_sheet(ods_sheets, keyword, excel_sheet_name):
    """Find the best matching ODS sheet for a given duty type keyword."""
    candidates = []
    for name in ods_sheets:
        # Skip external references
        if name.startswith("'file:"):
            continue
        if keyword in name:
            # Prefer 114-2, then 114-1, then others
            priority = 0
            if '114-2' in name:
                priority = 3
            elif '114-1' in name:
                priority = 2
            elif '113-2' in name:
                priority = 1
            candidates.append((priority, name))

    if not candidates:
        return None

    # Sort by priority descending, then by name length (shorter = more specific)
    candidates.sort(key=lambda x: (-x[0], len(x[1])))
    return candidates[0][1]


# =============================================================================
# Excel Parsing Functions
# =============================================================================

def get_excel_sheet_data(ws):
    """Extract structured data from an openpyxl worksheet."""
    data = {
        'rows': [],
        'merges': [],
        'col_widths': {},
        'row_heights': {},
    }

    # Column widths
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width is not None:
            data['col_widths'][col_letter] = dim.width

    # Row heights
    for row_num, dim in ws.row_dimensions.items():
        if dim.height is not None:
            data['row_heights'][row_num] = dim.height

    # Merged cells
    for merge in ws.merged_cells.ranges:
        data['merges'].append({
            'min_row': merge.min_row, 'max_row': merge.max_row,
            'min_col': merge.min_col, 'max_col': merge.max_col,
            'row_span': merge.max_row - merge.min_row + 1,
            'col_span': merge.max_col - merge.min_col + 1,
        })

    # Cell data
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        row_data = []
        for cell in row:
            cell_info = {
                'value': cell.value,
                'text': str(cell.value) if cell.value is not None else '',
                'row': cell.row,
                'col': cell.column,
                'font': None,
                'fill': None,
                'border': None,
                'alignment': None,
            }
            if cell.font:
                cell_info['font'] = {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'color': str(cell.font.color.rgb) if cell.font.color and cell.font.color.rgb else None,
                }
            if cell.fill:
                fg = cell.fill.fgColor
                cell_info['fill'] = {
                    'type': cell.fill.fill_type,
                    'fgColor': str(fg.rgb) if fg and fg.rgb and fg.rgb != '00000000' else None,
                }
            if cell.border:
                cell_info['border'] = {
                    'top': str(cell.border.top.style) if cell.border.top and cell.border.top.style else None,
                    'bottom': str(cell.border.bottom.style) if cell.border.bottom and cell.border.bottom.style else None,
                    'left': str(cell.border.left.style) if cell.border.left and cell.border.left.style else None,
                    'right': str(cell.border.right.style) if cell.border.right and cell.border.right.style else None,
                }
            if cell.alignment:
                cell_info['alignment'] = {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'wrapText': cell.alignment.wrapText,
                    'shrinkToFit': cell.alignment.shrinkToFit,
                }
            row_data.append(cell_info)
        data['rows'].append(row_data)

    return data


# =============================================================================
# Conversion Utilities
# =============================================================================

def cm_to_char_width(cm_str):
    """Convert ODS cm width to approximate Excel character width."""
    if not cm_str:
        return None
    m = re.match(r'([\d.]+)cm', cm_str)
    if m:
        cm = float(m.group(1))
        # 1 char width ~ 0.22 cm (approximate)
        return round(cm / 0.22, 2)
    m = re.match(r'([\d.]+)in', cm_str)
    if m:
        inches = float(m.group(1))
        return round(inches * 2.54 / 0.22, 2)
    return cm_str


def cm_to_pt(cm_str):
    """Convert ODS cm height to points."""
    if not cm_str:
        return None
    m = re.match(r'([\d.]+)cm', cm_str)
    if m:
        cm = float(m.group(1))
        return round(cm * 28.3465, 2)  # 1cm = 28.3465pt
    m = re.match(r'([\d.]+)in', cm_str)
    if m:
        inches = float(m.group(1))
        return round(inches * 72, 2)
    return cm_str


def parse_ods_font_size(size_str):
    """Parse ODS font size string to pt value."""
    if not size_str:
        return None
    m = re.match(r'([\d.]+)pt', size_str)
    if m:
        return float(m.group(1))
    return size_str


def parse_ods_border(border_str):
    """Parse ODS border string like '0.06pt solid #000000'."""
    if not border_str or border_str == 'none':
        return {'width': None, 'style': None, 'color': None}
    parts = border_str.strip().split()
    result = {'width': None, 'style': None, 'color': None}
    if len(parts) >= 1:
        result['width'] = parts[0]
    if len(parts) >= 2:
        result['style'] = parts[1]
    if len(parts) >= 3:
        result['color'] = parts[2]
    return result


def ods_border_to_excel_style(border_str):
    """Map ODS border width to Excel border style name."""
    if not border_str or border_str == 'none':
        return None
    parsed = parse_ods_border(border_str)
    width = parsed.get('width', '')
    if not width:
        return None
    m = re.match(r'([\d.]+)pt', width)
    if m:
        w = float(m.group(1))
        if w >= 2.5:
            return 'thick'
        elif w >= 1.5:
            return 'medium'
        elif w >= 0.5:
            return 'thin'
        else:
            return 'hair'
    return 'thin'


def repr_text(t):
    """Repr a text for comparison, showing control chars."""
    if t is None:
        return repr(None)
    return repr(t)


# =============================================================================
# Comparison Functions
# =============================================================================

class DiffCollector:
    def __init__(self):
        self.diffs = defaultdict(list)  # category -> list of diff descriptions
        self.counts = defaultdict(int)

    def add(self, category, msg):
        self.diffs[category].append(msg)
        self.counts[category] += 1


def identify_cell_role(row_idx, col_idx, text, total_rows):
    """Try to identify what role a cell plays: title, date, weekday, time_slot, data, sig, note."""
    if row_idx == 0:
        return 'title'
    if row_idx == 1:
        if col_idx == 0:
            return 'header_label'
        return 'date'
    if row_idx == 2:
        if col_idx == 0:
            return 'header_label'
        return 'weekday'
    if col_idx == 0 and text:
        return 'time_slot'
    if text and ('簽章' in str(text) or '備' in str(text) or '註' in str(text)):
        return 'note_or_sig'
    return 'data'


def compare_duty_type(excel_ws, ods_sheet_data, ods_styles, duty_name, ods_sheet_name, dc):
    """Compare a single duty type between Excel and ODS."""
    prefix = f"[{duty_name}]"
    xl = get_excel_sheet_data(excel_ws)

    ods = ods_sheet_data
    if not ods or not ods['rows']:
        dc.add('N_SheetNames', f"{prefix} ODS sheet '{ods_sheet_name}' has no data rows")
        return

    # =========================================================================
    # M. Sheet Names
    # =========================================================================
    dc.add('M_SheetNames', f"{prefix} Excel='{excel_ws.title}' vs ODS='{ods_sheet_name}' (expected mismatch, different naming convention)")

    # =========================================================================
    # Helper: get first N non-empty rows from both
    # =========================================================================
    xl_rows = xl['rows']
    ods_rows = ods['rows']

    max_compare_rows = min(len(xl_rows), len(ods_rows), 40)

    # =========================================================================
    # A. Title text format
    # =========================================================================
    # Title is typically row 0 (row 1 in Excel 1-based)
    if xl_rows and ods_rows:
        xl_title = ''
        for c in xl_rows[0]:
            if c['text']:
                xl_title = c['text']
                break
        ods_title = ''
        if ods_rows[0]['cells']:
            for c in ods_rows[0]['cells']:
                if c['text']:
                    ods_title = c['text']
                    break

        if xl_title != ods_title:
            dc.add('A_TitleText', f"{prefix} Title text differs:\n  Excel: {repr_text(xl_title)}\n  ODS:   {repr_text(ods_title)}")
        else:
            # Check for bracket width, spacing differences
            pass  # Exact match

    # =========================================================================
    # B. Date format + C. Weekday format
    # =========================================================================
    # Row 1 = dates, Row 2 = weekdays (0-indexed)
    if len(xl_rows) > 1 and len(ods_rows) > 1:
        # Collect date cells from Excel
        xl_dates = [c['text'] for c in xl_rows[1] if c['text']]
        ods_date_row = ods_rows[1] if len(ods_rows) > 1 else None
        ods_dates = []
        if ods_date_row:
            ods_dates = [c['text'] for c in ods_date_row['cells'] if c['text']]

        if xl_dates and ods_dates:
            # Compare first few date formats
            for i in range(min(3, len(xl_dates), len(ods_dates))):
                if xl_dates[i] != ods_dates[i]:
                    dc.add('B_DateFormat', f"{prefix} Date[{i}] differs:\n  Excel: {repr_text(xl_dates[i])}\n  ODS:   {repr_text(ods_dates[i])}")

    if len(xl_rows) > 2 and len(ods_rows) > 2:
        xl_weekdays = [c['text'] for c in xl_rows[2] if c['text']]
        ods_wd_row = ods_rows[2] if len(ods_rows) > 2 else None
        ods_weekdays = []
        if ods_wd_row:
            ods_weekdays = [c['text'] for c in ods_wd_row['cells'] if c['text']]

        if xl_weekdays and ods_weekdays:
            for i in range(min(3, len(xl_weekdays), len(ods_weekdays))):
                if xl_weekdays[i] != ods_weekdays[i]:
                    dc.add('C_WeekdayFormat', f"{prefix} Weekday[{i}] differs:\n  Excel: {repr_text(xl_weekdays[i])}\n  ODS:   {repr_text(ods_weekdays[i])}")

    # =========================================================================
    # D. Time slot format
    # =========================================================================
    # Time slots are in column A, starting from row 3
    xl_timeslots = []
    ods_timeslots = []
    for r_idx in range(3, min(len(xl_rows), 30)):
        if xl_rows[r_idx] and xl_rows[r_idx][0]['text']:
            xl_timeslots.append((r_idx, xl_rows[r_idx][0]['text']))
    for r_idx in range(3, min(len(ods_rows), 30)):
        if ods_rows[r_idx]['cells'] and ods_rows[r_idx]['cells'][0]['text']:
            ods_timeslots.append((r_idx, ods_rows[r_idx]['cells'][0]['text']))

    for i in range(min(len(xl_timeslots), len(ods_timeslots))):
        xl_r, xl_t = xl_timeslots[i]
        ods_r, ods_t = ods_timeslots[i]
        if xl_t != ods_t:
            dc.add('D_TimeSlotFormat', f"{prefix} TimeSlot[{i}] (xlRow{xl_r+1}/odsRow{ods_r}) differs:\n  Excel: {repr_text(xl_t)}\n  ODS:   {repr_text(ods_t)}")

    if len(xl_timeslots) != len(ods_timeslots):
        dc.add('D_TimeSlotFormat', f"{prefix} Number of time slots differs: Excel={len(xl_timeslots)}, ODS={len(ods_timeslots)}")

    # =========================================================================
    # E. Font sizes + F. Font names (cell by cell for first ~6 rows)
    # =========================================================================
    for r_idx in range(min(max_compare_rows, 8)):
        if r_idx >= len(xl_rows) or r_idx >= len(ods_rows):
            break

        xl_row = xl_rows[r_idx]
        ods_cells = ods_rows[r_idx]['cells']

        for c_idx in range(min(len(xl_row), len(ods_cells), 15)):
            xl_cell = xl_row[c_idx]
            ods_cell = ods_cells[c_idx]

            role = identify_cell_role(r_idx, c_idx, xl_cell['text'] or ods_cell['text'], len(xl_rows))

            # Font size
            xl_font_size = xl_cell['font']['size'] if xl_cell['font'] else None
            ods_style = ods_cell.get('style_data', {})
            ods_font_size_raw = ods_style.get('font-size', '') or ods_style.get('font-size-asian', '')
            ods_font_size = parse_ods_font_size(ods_font_size_raw)

            if xl_font_size and ods_font_size and xl_font_size != ods_font_size:
                dc.add('E_FontSizes', f"{prefix} Row{r_idx+1} Col{c_idx+1} ({role}): Excel={xl_font_size}pt, ODS={ods_font_size}pt")

            # Font name
            xl_font_name = xl_cell['font']['name'] if xl_cell['font'] else None
            ods_font_name = ods_style.get('font-name', '') or ods_style.get('font-name-asian', '')

            if xl_font_name and ods_font_name:
                # Normalize: DFKai-SB and 標楷體 are equivalent
                xl_fn_norm = xl_font_name.lower().replace(' ', '')
                ods_fn_norm = ods_font_name.lower().replace(' ', '')
                equiv_sets = [
                    {'dfkai-sb', '標楷體', '標楷體-繁'},
                ]
                matched = False
                for eq_set in equiv_sets:
                    if xl_fn_norm in eq_set and ods_fn_norm in eq_set:
                        matched = True
                        break
                if not matched and xl_font_name != ods_font_name:
                    dc.add('F_FontNames', f"{prefix} Row{r_idx+1} Col{c_idx+1} ({role}): Excel='{xl_font_name}', ODS='{ods_font_name}'")

    # =========================================================================
    # G. Border styles
    # =========================================================================
    # Build a set of cells that are inside a merged range but not the anchor
    ods_merged_non_anchor = set()
    for mg in ods['merges']:
        for r in range(mg['row'], mg['row'] + mg['row_span']):
            for c in range(mg['col'], mg['col'] + mg['col_span']):
                if r != mg['row'] or c != mg['col']:
                    ods_merged_non_anchor.add((r, c))

    # Similarly for Excel
    xl_merged_non_anchor = set()
    for mg in xl['merges']:
        for r in range(mg['min_row'], mg['max_row'] + 1):
            for c in range(mg['min_col'], mg['max_col'] + 1):
                if r != mg['min_row'] or c != mg['min_col']:
                    xl_merged_non_anchor.add((r, c))

    # Count border diffs by type for summary
    border_diff_summary = defaultdict(int)
    for r_idx in range(min(max_compare_rows, 10)):
        if r_idx >= len(xl_rows) or r_idx >= len(ods_rows):
            break
        xl_row = xl_rows[r_idx]
        ods_cells = ods_rows[r_idx]['cells']

        for c_idx in range(min(len(xl_row), len(ods_cells), 15)):
            # Skip non-anchor cells in merged regions (ODS doesn't store borders there)
            if (r_idx, c_idx) in ods_merged_non_anchor:
                continue

            xl_cell = xl_row[c_idx]
            ods_cell = ods_cells[c_idx]
            role = identify_cell_role(r_idx, c_idx, xl_cell['text'] or ods_cell['text'], len(xl_rows))

            ods_style = ods_cell.get('style_data', {})

            # Get ODS border info
            ods_border_all = ods_style.get('border', '')
            ods_borders = {
                'top': ods_style.get('border-top', '') or ods_border_all,
                'bottom': ods_style.get('border-bottom', '') or ods_border_all,
                'left': ods_style.get('border-left', '') or ods_border_all,
                'right': ods_style.get('border-right', '') or ods_border_all,
            }

            xl_border = xl_cell.get('border', {}) or {}

            for side in ['top', 'bottom', 'left', 'right']:
                xl_b = xl_border.get(side) if xl_border else None
                ods_b_str = ods_borders.get(side, '')
                ods_b = ods_border_to_excel_style(ods_b_str)

                if xl_b != ods_b:
                    # Only report if at least one has a border
                    if xl_b or ods_b:
                        diff_type = f"Excel={xl_b} vs ODS={ods_b}"
                        border_diff_summary[diff_type] += 1
                        # Only print first 5 per duty type
                        if dc.counts.get('G_BorderStyles', 0) < 20 or border_diff_summary[diff_type] <= 2:
                            dc.add('G_BorderStyles', f"{prefix} Row{r_idx+1} Col{c_idx+1} ({role}) {side}: Excel='{xl_b}', ODS='{ods_b}' (raw: {repr(ods_b_str)})")
                        else:
                            dc.counts['G_BorderStyles'] = dc.counts.get('G_BorderStyles', 0) + 1

    if border_diff_summary:
        dc.add('G_BorderStyles', f"{prefix} Border diff summary: {dict(border_diff_summary)}")

    # =========================================================================
    # H. Fill colors
    # =========================================================================
    for r_idx in range(min(max_compare_rows, 10)):
        if r_idx >= len(xl_rows) or r_idx >= len(ods_rows):
            break
        xl_row = xl_rows[r_idx]
        ods_cells = ods_rows[r_idx]['cells']

        for c_idx in range(min(len(xl_row), len(ods_cells), 15)):
            xl_cell = xl_row[c_idx]
            ods_cell = ods_cells[c_idx]
            role = identify_cell_role(r_idx, c_idx, xl_cell['text'] or ods_cell['text'], len(xl_rows))

            ods_style = ods_cell.get('style_data', {})
            ods_bg = ods_style.get('background-color', '')
            xl_fill = xl_cell.get('fill', {}) or {}
            xl_fg = xl_fill.get('fgColor')

            # Normalize: treat transparent, #FFFFFF, FFFFFFFF all as "no fill"
            def norm_color(c):
                if not c or c in ('transparent', 'None', '000000', '00000000'):
                    return None
                c = c.upper().replace('#', '')
                if len(c) == 8:
                    c = c[2:]  # strip alpha
                if c in ('FFFFFF', '000000'):
                    return None
                return c

            ods_bg_norm = norm_color(ods_bg)
            xl_fg_norm = norm_color(xl_fg)

            if ods_bg_norm != xl_fg_norm:
                if ods_bg_norm or xl_fg_norm:
                    dc.add('H_FillColors', f"{prefix} Row{r_idx+1} Col{c_idx+1} ({role}): Excel='{xl_fg}' -> {xl_fg_norm}, ODS='{ods_bg}' -> {ods_bg_norm}")

    # =========================================================================
    # I. Alignment
    # =========================================================================
    for r_idx in range(min(max_compare_rows, 10)):
        if r_idx >= len(xl_rows) or r_idx >= len(ods_rows):
            break
        xl_row = xl_rows[r_idx]
        ods_cells = ods_rows[r_idx]['cells']

        for c_idx in range(min(len(xl_row), len(ods_cells), 15)):
            xl_cell = xl_row[c_idx]
            ods_cell = ods_cells[c_idx]
            role = identify_cell_role(r_idx, c_idx, xl_cell['text'] or ods_cell['text'], len(xl_rows))

            ods_style = ods_cell.get('style_data', {})
            xl_align = xl_cell.get('alignment', {}) or {}

            # Horizontal alignment
            ods_h_align = ods_style.get('text-align', '')
            xl_h_align = xl_align.get('horizontal', '') or ''

            # Map ODS alignment values to Excel
            ods_h_map = {'center': 'center', 'start': 'left', 'end': 'right', 'left': 'left', 'right': 'right'}
            ods_h_mapped = ods_h_map.get(ods_h_align, ods_h_align)

            if ods_h_mapped and xl_h_align and ods_h_mapped != xl_h_align:
                dc.add('I_Alignment', f"{prefix} Row{r_idx+1} Col{c_idx+1} ({role}) h-align: Excel='{xl_h_align}', ODS='{ods_h_mapped}' (raw: {ods_h_align})")

            # Vertical alignment
            ods_v_align = ods_style.get('vertical-align', '')
            xl_v_align = xl_align.get('vertical', '') or ''

            ods_v_map = {'middle': 'center', 'top': 'top', 'bottom': 'bottom'}
            ods_v_mapped = ods_v_map.get(ods_v_align, ods_v_align)

            if ods_v_mapped and xl_v_align and ods_v_mapped != xl_v_align:
                dc.add('I_Alignment', f"{prefix} Row{r_idx+1} Col{c_idx+1} ({role}) v-align: Excel='{xl_v_align}', ODS='{ods_v_mapped}' (raw: {ods_v_align})")

            # Wrap text
            ods_wrap = ods_style.get('wrap-option', '')
            xl_wrap = xl_align.get('wrapText')
            ods_wrap_bool = ods_wrap == 'wrap' if ods_wrap else None

            if ods_wrap_bool is not None and xl_wrap is not None and ods_wrap_bool != xl_wrap:
                dc.add('I_Alignment', f"{prefix} Row{r_idx+1} Col{c_idx+1} ({role}) wrapText: Excel={xl_wrap}, ODS={ods_wrap_bool} (raw: {ods_wrap})")

    # =========================================================================
    # J. Row heights
    # =========================================================================
    xl_heights = xl['row_heights']
    for r_idx in range(min(max_compare_rows, 20)):
        if r_idx >= len(ods_rows):
            break
        ods_h = ods_rows[r_idx]['height']
        ods_h_pt = cm_to_pt(ods_h)
        xl_h = xl_heights.get(r_idx + 1)  # Excel is 1-based

        if ods_h_pt and xl_h:
            # Allow 2pt tolerance
            if isinstance(ods_h_pt, (int, float)) and isinstance(xl_h, (int, float)):
                if abs(ods_h_pt - xl_h) > 2:
                    role = 'title' if r_idx == 0 else 'date' if r_idx == 1 else 'weekday' if r_idx == 2 else f'data_row{r_idx}'
                    # Get text hint
                    text_hint = ''
                    if xl_rows and r_idx < len(xl_rows) and xl_rows[r_idx]:
                        for c in xl_rows[r_idx]:
                            if c['text']:
                                text_hint = c['text'][:20]
                                break
                    dc.add('J_RowHeights', f"{prefix} Row{r_idx+1} ({role}, '{text_hint}'): Excel={xl_h}pt, ODS={ods_h_pt}pt (raw: {ods_h})")

    # =========================================================================
    # K. Column widths
    # =========================================================================
    xl_widths = xl['col_widths']
    for c_idx in range(min(len(ods['columns']), 20)):
        col_info = ods['columns'][c_idx]
        ods_w_raw = col_info.get('width', '') or ''
        ods_w_char = cm_to_char_width(ods_w_raw)
        col_letter = get_column_letter(c_idx + 1)
        xl_w = xl_widths.get(col_letter)

        if ods_w_char and xl_w:
            if isinstance(ods_w_char, (int, float)) and isinstance(xl_w, (int, float)):
                if abs(ods_w_char - xl_w) > 2:  # Allow 2 char tolerance
                    dc.add('K_ColumnWidths', f"{prefix} Col {col_letter} (idx {c_idx}): Excel={xl_w:.2f} chars, ODS~={ods_w_char:.2f} chars (raw: {ods_w_raw})")

    # =========================================================================
    # L. Merge structure
    # =========================================================================
    xl_merges = xl['merges']
    ods_merges = ods['merges']

    # Compare merge counts
    if len(xl_merges) != len(ods_merges):
        dc.add('L_MergeStructure', f"{prefix} Total merges: Excel={len(xl_merges)}, ODS={len(ods_merges)}")

    # Compare title merge (row 0/1)
    xl_title_merges = [m for m in xl_merges if m['min_row'] == 1]
    ods_title_merges = [m for m in ods_merges if m['row'] == 0]
    if xl_title_merges and ods_title_merges:
        xl_tm = xl_title_merges[0]
        ods_tm = ods_title_merges[0]
        xl_span = f"cols {xl_tm['min_col']}-{xl_tm['max_col']} (span={xl_tm['col_span']})"
        ods_span = f"cols {ods_tm['col']}-{ods_tm['col']+ods_tm['col_span']-1} (span={ods_tm['col_span']})"
        if xl_tm['col_span'] != ods_tm['col_span']:
            dc.add('L_MergeStructure', f"{prefix} Title merge: Excel={xl_span}, ODS={ods_span}")

    # Compare date/weekday merges
    xl_date_merges = sorted([m for m in xl_merges if m['min_row'] == 2], key=lambda m: m['min_col'])
    ods_date_merges = sorted([m for m in ods_merges if m['row'] == 1], key=lambda m: m['col'])
    if len(xl_date_merges) != len(ods_date_merges):
        dc.add('L_MergeStructure', f"{prefix} Date row merges count: Excel={len(xl_date_merges)}, ODS={len(ods_date_merges)}")
    else:
        for i in range(min(len(xl_date_merges), len(ods_date_merges))):
            if xl_date_merges[i]['col_span'] != ods_date_merges[i]['col_span']:
                dc.add('L_MergeStructure', f"{prefix} Date merge[{i}] col_span: Excel={xl_date_merges[i]['col_span']}, ODS={ods_date_merges[i]['col_span']}")

    # Time slot merges (col A, multiple rows)
    xl_time_merges = sorted([m for m in xl_merges if m['min_col'] == 1 and m['min_row'] > 2], key=lambda m: m['min_row'])
    ods_time_merges = sorted([m for m in ods_merges if m['col'] == 0 and m['row'] > 2], key=lambda m: m['row'])
    if len(xl_time_merges) != len(ods_time_merges):
        dc.add('L_MergeStructure', f"{prefix} Time slot merges count: Excel={len(xl_time_merges)}, ODS={len(ods_time_merges)}")

    # =========================================================================
    # N. Note content
    # =========================================================================
    # Look for 備註/備考/注意 in both
    xl_notes = []
    ods_notes = []
    for r_idx, row in enumerate(xl_rows):
        for cell in row:
            t = str(cell['text'])
            if '備註' in t or '備考' in t or '注意' in t or '附註' in t:
                xl_notes.append((r_idx, cell['col'], t))
            elif cell['text'] and r_idx >= len(xl_rows) - 5:  # Last 5 rows likely notes
                xl_notes.append((r_idx, cell['col'], t))

    for r_idx in range(max(0, len(ods_rows) - 10), len(ods_rows)):
        for cell in ods_rows[r_idx]['cells']:
            t = cell['text']
            if t and ('備註' in t or '備考' in t or '注意' in t or '附註' in t):
                ods_notes.append((r_idx, cell.get('col_idx', 0), t))

    if xl_notes or ods_notes:
        xl_note_texts = [n[2] for n in xl_notes[:5]]
        ods_note_texts = [n[2] for n in ods_notes[:5]]
        if xl_note_texts != ods_note_texts:
            dc.add('N_NoteContent', f"{prefix} Note texts differ:\n  Excel({len(xl_note_texts)}): {[repr_text(t)[:80] for t in xl_note_texts]}\n  ODS({len(ods_note_texts)}): {[repr_text(t)[:80] for t in ods_note_texts]}")

    # =========================================================================
    # O. Data structure
    # =========================================================================
    # Count time slots, data rows, total rows/cols
    xl_total_rows = len(xl_rows)
    xl_total_cols = max(len(r) for r in xl_rows) if xl_rows else 0
    ods_total_rows = len(ods_rows)
    ods_total_cols = max(len(r['cells']) for r in ods_rows) if ods_rows else 0

    if xl_total_rows != ods_total_rows:
        dc.add('O_DataStructure', f"{prefix} Total rows: Excel={xl_total_rows}, ODS={ods_total_rows}")
    # Don't compare total cols directly as ODS often has many empty trailing columns


def compare_all():
    """Main comparison function."""
    print("=" * 80)
    print("EXHAUSTIVE FORMAT COMPARISON: Exported Excel vs Reference ODS")
    print("=" * 80)
    print(f"\nExcel: {EXCEL_PATH}")
    print(f"ODS:   {ODS_PATH}")
    print()

    # Load files
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print(f"  Sheets: {wb.sheetnames}")

    print("Loading ODS file...")
    ods_sheets, ods_styles = parse_ods(ODS_PATH)
    ods_sheet_names = [n for n in ods_sheets.keys() if not n.startswith("'file:")]
    print(f"  Local sheets: {len(ods_sheet_names)}")

    dc = DiffCollector()

    # Map each Excel sheet to the best ODS sheet
    print("\n" + "=" * 80)
    print("SHEET MAPPING")
    print("=" * 80)

    sheet_mapping = {}
    for xl_name, keyword in KEYWORD_MAP.items():
        ods_name = find_best_ods_sheet(ods_sheets, keyword, xl_name)
        sheet_mapping[xl_name] = ods_name
        status = "FOUND" if ods_name else "NOT FOUND"
        print(f"  {xl_name} ({keyword}) -> {ods_name} [{status}]")

    # Compare each duty type
    for xl_name, ods_name in sheet_mapping.items():
        print(f"\n{'='*80}")
        print(f"COMPARING: {xl_name} vs {ods_name}")
        print(f"{'='*80}")

        if xl_name not in wb.sheetnames:
            dc.add('M_SheetNames', f"Excel sheet '{xl_name}' not found")
            continue
        if not ods_name or ods_name not in ods_sheets:
            dc.add('M_SheetNames', f"ODS sheet for '{xl_name}' not found (keyword: {KEYWORD_MAP[xl_name]})")
            continue

        xl_ws = wb[xl_name]
        ods_data = ods_sheets[ods_name]

        compare_duty_type(xl_ws, ods_data, ods_styles, xl_name, ods_name, dc)

    # =========================================================================
    # Print Results
    # =========================================================================
    print("\n" + "=" * 80)
    print("DIFFERENCES FOUND (only mismatches shown)")
    print("=" * 80)

    categories = [
        ('A_TitleText', 'A. Title Text Format'),
        ('B_DateFormat', 'B. Date Format'),
        ('C_WeekdayFormat', 'C. Weekday Format'),
        ('D_TimeSlotFormat', 'D. Time Slot Format'),
        ('E_FontSizes', 'E. Font Sizes'),
        ('F_FontNames', 'F. Font Names'),
        ('G_BorderStyles', 'G. Border Styles'),
        ('H_FillColors', 'H. Fill Colors'),
        ('I_Alignment', 'I. Alignment'),
        ('J_RowHeights', 'J. Row Heights'),
        ('K_ColumnWidths', 'K. Column Widths'),
        ('L_MergeStructure', 'L. Merge Structure'),
        ('M_SheetNames', 'M. Sheet Names'),
        ('N_NoteContent', 'N. Note Content'),
        ('O_DataStructure', 'O. Data Structure'),
    ]

    total_diffs = 0
    for key, label in categories:
        diffs = dc.diffs.get(key, [])
        if diffs:
            print(f"\n--- {label} ({len(diffs)} differences) ---")
            for d in diffs[:30]:  # Limit output per category
                print(f"  {d}")
            if len(diffs) > 30:
                print(f"  ... and {len(diffs) - 30} more")
            total_diffs += len(diffs)
        else:
            print(f"\n--- {label}: ALL MATCH (0 differences) ---")

    # =========================================================================
    # Summary
    # =========================================================================
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"\n{'Category':<35} {'Differences':>12}")
    print("-" * 50)
    for key, label in categories:
        count = dc.counts.get(key, 0)
        marker = " !!!" if count > 0 else ""
        print(f"  {label:<33} {count:>8}{marker}")
    print("-" * 50)
    print(f"  {'TOTAL':<33} {total_diffs:>8}")

    print(f"\n{'='*80}")
    if total_diffs == 0:
        print("VERDICT: PERFECT MATCH -- All formats are identical.")
    elif total_diffs <= 10:
        print(f"VERDICT: NEAR MATCH -- Only {total_diffs} minor differences found.")
    elif total_diffs <= 50:
        print(f"VERDICT: PARTIAL MATCH -- {total_diffs} differences found, some formatting adjustments needed.")
    else:
        print(f"VERDICT: SIGNIFICANT DIFFERENCES -- {total_diffs} differences found, substantial formatting work needed.")
    print("=" * 80)


if __name__ == '__main__':
    compare_all()
